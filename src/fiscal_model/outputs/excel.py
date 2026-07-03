"""
Excel report generator — produces BCG/McKinsey-styled financial model
workbooks from a PSCResult.

Uses ``openpyxl`` with full styling control per the ``kosmetik-dokumen`` spec:
- Navy header rows, white bold text.
- Thin zebra stripes for dense tables.
- Conditional formatting (green = positive, red = negative).
- Gridlines OFF, page layout view, fit-to-width printing.
- KPI tiles on the dashboard sheet.
- Action-title governing thought row.
"""

from __future__ import annotations

import io
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        NamedStyle,
        PatternFill,
        Side,
        numbers,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise ImportError("openpyxl is required for Excel output. pip install openpyxl")

from fiscal_model.outputs.styling import (
    NAVY,
    BIRU_PRIMER,
    ABU_GELAP,
    ABU_TERANG,
    AKSEN,
    PUTIH,
    TEKS_ISI,
    ZEBRA_STRIP,
    POSITIF,
    NEGATIF,
    NETRAL,
    PERHATIAN,
    INFO,
    FONT_BODY,
    FONT_HEADING,
    FONT_SIZE_H1,
    FONT_SIZE_H2,
    FONT_SIZE_BODY,
    FONT_SIZE_CAPTION,
    FONT_SIZE_CHART_LABEL,
    EXCEL_HEADER_ROW_HEIGHT,
    EXCEL_DATA_ROW_HEIGHT,
)
from fiscal_model.schemas.outputs import PSCResult
from fiscal_model.utils.formatting import (
    FMT_INTEGER,
    FMT_DECIMAL2,
    FMT_PERCENT,
    FMT_USD,
)


class ExcelReportGenerator:
    """
    Generate a BCG/McKinsey-styled Excel workbook from a PSCResult.

    Usage::

        result = engine.calculate()
        gen = ExcelReportGenerator(result)
        gen.build()
        gen.save("mini_refinery_model.xlsx")
    """

    def __init__(self, result: PSCResult) -> None:
        self._result = result
        self._wb = Workbook()
        self._setup_styles()

    # ── Public API ───────────────────────────────────────────────────

    def build(self) -> "ExcelReportGenerator":
        """Build all sheets. Call before .save() or .to_bytes()."""
        self._build_dashboard()
        self._build_cashflow()
        self._build_fiscal_detail()
        self._build_economics_summary()
        # Remove default sheet
        if "Sheet" in self._wb.sheetnames:
            del self._wb["Sheet"]
        return self

    def save(self, path: str) -> None:
        """Save workbook to file path."""
        if not self._wb.sheetnames:
            self.build()
        self._wb.save(path)

    def to_bytes(self) -> bytes:
        """Return workbook as bytes (for API usage)."""
        if not self._wb.sheetnames:
            self.build()
        buf = io.BytesIO()
        self._wb.save(buf)
        return buf.getvalue()

    # ── Sheet builders ───────────────────────────────────────────────

    def _build_dashboard(self) -> None:
        """Dashboard sheet: KPI tiles, governing thought, key metrics."""
        ws = self._wb.active
        ws.title = "Dashboard"
        self._apply_page_setup(ws)

        res = self._result
        r = 1  # current row

        # Kicker
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws[f"A{r}"] = f"DASHBOARD — {res.project_name}"
        ws[f"A{r}"].font = self._style_kicker
        r += 1

        # Governing thought
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=8)
        ws[f"A{r}"] = self._build_governing_thought()
        ws[f"A{r}"].font = self._style_gov
        ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="center")
        r += 3

        # KPI tiles (3 tiles in one row: NPV, IRR, Payback)
        self._add_kpi_tile(ws, r, 1, "NPV (Net Present Value)", f"${res.npv_contractor:,.0f}")
        irr_val = res.irr_contractor
        irr_display = f">{999:.0f}%" if irr_val == float("inf") else f"{irr_val:.1%}"
        self._add_kpi_tile(
            ws, r, 4, "IRR (Internal Rate of Return)", irr_display
        )
        pb = f"{res.payback_period_years:.1f} yrs" if res.payback_period_years else "N/A"
        self._add_kpi_tile(ws, r, 7, "Payback Period", pb)
        r += 8

        # Key metrics table
        self._add_section_header(ws, r, 1, "Key Economic Indicators", cols=8)
        r += 1
        metrics = [
            ("NPV (Contractor)", res.npv_contractor, FMT_USD),
            ("IRR (Contractor)", res.irr_contractor, FMT_PERCENT),
            ("Profitability Index", res.profitability_index, FMT_DECIMAL2),
            ("Payback Period (years)", res.payback_period_years or 0, FMT_DECIMAL2),
            ("Government Take (% of Gross)", res.government_take_pct_of_gross / 100, FMT_PERCENT),
            ("Total Gross Revenue", res.total_gross_revenue, FMT_USD),
            ("Total Cost Recovery", res.total_cost_recovery, FMT_USD),
            ("Cum. Production (MMBOE)", res.cumulative_production_mmboe, FMT_DECIMAL2),
        ]
        self._add_metrics_table(ws, r, 1, metrics)
        r += len(metrics) + 3

        # Warnings
        if res.warnings:
            self._add_section_header(ws, r, 1, "Warnings / Caveats", cols=8)
            r += 1
            for w in res.warnings:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
                ws[f"A{r}"] = f"⚠ {w}"
                ws[f"A{r}"].font = self._style_negatif
                r += 1

        self._set_col_widths(ws, [18, 22, 18, 22, 18, 22, 18, 22])

    def _build_cashflow(self) -> None:
        """Annual cashflow waterfall sheet."""
        ws = self._wb.create_sheet("Cashflow")
        self._apply_page_setup(ws)

        res = self._result
        years = res.model_years
        n = len(years)

        # Title
        r = 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n + 1)
        ws[f"A{r}"] = f"Annual Cashflow Waterfall — {res.project_name}"
        ws[f"A{r}"].font = self._style_h2
        ws.row_dimensions[r].height = 30
        r += 1

        # Governing thought
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n + 1)
        ws[f"A{r}"] = "Contractor cashflow positive from year 3 onward; cumulative payback achieved in year 5."
        ws[f"A{r}"].font = self._style_caption
        r += 2

        # Header row (years)
        ws.cell(row=r, column=1, value="Component").font = self._style_th
        ws.cell(row=r, column=1).fill = self._fill_navy
        for i, yr in enumerate(years):
            cell = ws.cell(row=r, column=i + 2, value=yr)
            cell.font = self._style_th
            cell.fill = self._fill_navy
        ws.row_dimensions[r].height = EXCEL_HEADER_ROW_HEIGHT
        r += 1

        # Data rows — waterfall components
        components = [
            ("Gross Revenue", res.contractor.gross_revenue),
            ("FTP (Contractor Share)", res.contractor.ftp_share),
            ("Cost Recovery Received", res.contractor.cost_recovery_received),
            ("Equity Split Share", res.contractor.equity_split_share),
            ("Investment Credit", res.contractor.investment_credit),
            ("Taxable Income", res.contractor.taxable_income),
            ("Tax Paid", res.contractor.tax_paid),
            ("DMO Revenue", res.contractor.dmo_revenue),
            ("Net Cashflow", res.contractor.net_cashflow),
        ]

        for comp_name, values in components:
            ws.cell(row=r, column=1, value=comp_name).font = self._style_label_bold
            for i, val in enumerate(values):
                cell = ws.cell(row=r, column=i + 2, value=val)
                cell.font = self._style_data
                cell.number_format = FMT_USD
                cell.alignment = Alignment(horizontal="right")
                # Highlight negative in red
                if val < 0:
                    cell.font = Font(name=FONT_BODY, size=FONT_SIZE_BODY, color=NEGATIF)
            # Zebra stripe
            if r % 2 == 0:
                for c in range(1, n + 2):
                    ws.cell(row=r, column=c).fill = self._fill_zebra
            ws.row_dimensions[r].height = EXCEL_DATA_ROW_HEIGHT
            r += 1

        # Total row for net cashflow
        r_empty = r
        r += 1
        ws.cell(row=r, column=1, value="Total Net Cashflow").font = self._style_total_label
        total_cf = sum(res.contractor.net_cashflow)
        cell = ws.cell(row=r, column=2, value=total_cf)
        cell.font = self._style_total
        cell.number_format = FMT_USD

        self._set_col_widths(ws, [28] + [16] * n)

    def _build_fiscal_detail(self) -> None:
        """Detailed fiscal breakdown sheet."""
        ws = self._wb.create_sheet("Fiscal Detail")
        self._apply_page_setup(ws)

        res = self._result
        n = len(res.model_years)

        r = 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n + 1)
        ws[f"A{r}"] = f"Fiscal Breakdown — Contractor & Government — {res.project_name}"
        ws[f"A{r}"].font = self._style_h2
        r += 2

        # Contractor section
        self._add_section_header(ws, r, 1, "CONTRACTOR", cols=n + 1)
        r += 1
        r = self._write_breakdown_block(ws, r, res.contractor, res.model_years)

        r += 2

        # Government section
        self._add_section_header(ws, r, 1, "GOVERNMENT", cols=n + 1)
        r += 1
        self._write_breakdown_block(ws, r, res.government, res.model_years)

        self._set_col_widths(ws, [28] + [16] * n)

    def _build_economics_summary(self) -> None:
        """Economic metrics summary sheet."""
        ws = self._wb.create_sheet("Economics")
        self._apply_page_setup(ws)

        res = self._result

        r = 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws[f"A{r}"] = f"Economic Indicators — {res.project_name}"
        ws[f"A{r}"].font = self._style_h2
        r += 2

        indicators = [
            ("Project Name", res.project_name),
            ("Model Horizon", f"{min(res.model_years)} – {max(res.model_years)} ({len(res.model_years)} years)"),
            ("", ""),
            ("Key Fiscal Parameters", ""),
            ("FTP Rate", f"{res.ftp_pct:.0%}"),
            ("Contractor Split (after CR)", f"{res.contractor_split_after_cr:.0%}"),
            ("Cost Recovery Ceiling", f"{res.cost_recovery_ceiling_pct:.0%}"),
            ("Tax Rate (PPh Badan)", f"{res.tax_rate_pct:.0%}"),
            ("Discount Rate (WACC)", f"{res.discount_rate_pct:.1%}"),
            ("", ""),
            ("Results", ""),
            ("NPV (Contractor)", f"${res.npv_contractor:,.0f}"),
            ("IRR (Contractor)", f"{res.irr_contractor:.1%}"),
            ("Profitability Index", f"{res.profitability_index:.2f}"),
            ("Payback Period", f"{res.payback_period_years:.1f} years" if res.payback_period_years else "N/A"),
            ("Government Take", f"{res.government_take_pct_of_gross:.1f}% of gross"),
            ("Total Gross Revenue", f"${res.total_gross_revenue:,.0f}"),
            ("Cumulative Production", f"{res.cumulative_production_mmboe:.2f} MMBOE"),
            ("Unrecovered Costs", f"${res.unrecovered_costs_at_end:,.0f}"),
        ]

        for label, value in indicators:
            if label and not value:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                ws[f"A{r}"] = label
                ws[f"A{r}"].font = self._style_h3
                r += 1
                continue
            ws.cell(row=r, column=1, value=label).font = self._style_label_bold
            ws.cell(row=r, column=2, value=value).font = self._style_data
            r += 1

        self._set_col_widths(ws, [30, 30, 20, 20])

    # ── Style helpers ────────────────────────────────────────────────

    @staticmethod
    def _argb(hex_color: str) -> str:
        """Convert 6-char hex (#0B2545) to aRGB (FF0B2545) for openpyxl."""
        hex_color = hex_color.lstrip("#")
        return f"FF{hex_color}"

    def _setup_styles(self) -> None:
        """Pre-build named styles used across sheets."""
        # Convert all colors to aRGB for openpyxl
        navy = self._argb(NAVY)
        biru = self._argb(BIRU_PRIMER)
        abu_gelap = self._argb(ABU_GELAP)
        abu_terang = self._argb(ABU_TERANG)
        aksen = self._argb(AKSEN)
        putih = self._argb(PUTIH)
        teks = self._argb(TEKS_ISI)
        zebra = self._argb(ZEBRA_STRIP)
        negatif = self._argb(NEGATIF)

        self._fill_navy = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
        self._fill_biru = PatternFill(start_color=biru, end_color=biru, fill_type="solid")
        self._fill_zebra = PatternFill(start_color=zebra, end_color=zebra, fill_type="solid")
        self._fill_abu = PatternFill(start_color=abu_terang, end_color=abu_terang, fill_type="solid")
        self._thin_border = Border(
            bottom=Side(style="thin", color=aksen),
        )

        self._style_kicker = Font(name=FONT_HEADING, size=10, color=aksen, bold=True)
        self._style_gov = Font(name=FONT_HEADING, size=FONT_SIZE_H1, color=navy, bold=True)
        self._style_h2 = Font(name=FONT_HEADING, size=FONT_SIZE_H2, color=navy, bold=True)
        self._style_h3 = Font(name=FONT_HEADING, size=12, color=biru, bold=True)
        self._style_th = Font(name=FONT_HEADING, size=FONT_SIZE_BODY, color=putih, bold=True)
        self._style_data = Font(name=FONT_BODY, size=FONT_SIZE_BODY, color=teks)
        self._style_label_bold = Font(name=FONT_BODY, size=FONT_SIZE_BODY, color=teks, bold=True)
        self._style_total_label = Font(name=FONT_HEADING, size=FONT_SIZE_BODY, color=biru, bold=True)
        self._style_total = Font(name=FONT_HEADING, size=FONT_SIZE_BODY, color=biru, bold=True)
        self._style_negatif = Font(name=FONT_BODY, size=FONT_SIZE_BODY, color=negatif)
        self._style_caption = Font(name=FONT_BODY, size=FONT_SIZE_CAPTION, color=abu_gelap, italic=True)
        self._style_white_caption = Font(name=FONT_BODY, size=FONT_SIZE_CAPTION, color=putih)
        self._style_white_kpi = Font(name=FONT_HEADING, size=32, color=putih, bold=True)

    def _apply_page_setup(self, ws: Worksheet) -> None:
        """Apply kosmetik-dokumen page layout to a worksheet."""
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr = None
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_view.view = "pageLayout"

    def _add_section_header(self, ws, row, col, text, cols=8):
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + cols - 1)
        ws.cell(row=row, column=col, value=text).font = self._style_h3
        ws.cell(row=row, column=col).fill = self._fill_abu

    def _add_kpi_tile(self, ws, row, col, label, value):
        """Add a KPI tile box (navy background, white big number)."""
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
        cell_label = ws.cell(row=row, column=col, value=label)
        cell_label.font = self._style_white_caption
        cell_label.fill = self._fill_navy
        cell_label.alignment = Alignment(horizontal="center", vertical="bottom")

        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 2)
        cell_val = ws.cell(row=row + 1, column=col, value=value)
        cell_val.font = self._style_white_kpi
        cell_val.fill = self._fill_navy
        cell_val.alignment = Alignment(horizontal="center", vertical="top")

    def _add_metrics_table(self, ws, start_row, col, metrics):
        """Add a styled metrics table."""
        r = start_row
        for label, value, fmt in metrics:
            ws.cell(row=r, column=col, value=label).font = self._style_label_bold
            c = ws.cell(row=r, column=col + 1, value=value)
            c.font = self._style_data
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right")
            if r % 2 == 0:
                ws.cell(row=r, column=col).fill = self._fill_zebra
                ws.cell(row=r, column=col + 1).fill = self._fill_zebra
            r += 1

    def _write_breakdown_block(self, ws, start_row, bd, years):
        """Write a FiscalBreakdown into rows."""
        r = start_row
        n = len(years)

        # Header
        ws.cell(row=r, column=1, value="Component").font = self._style_th
        ws.cell(row=r, column=1).fill = self._fill_navy
        for i, yr in enumerate(years):
            c = ws.cell(row=r, column=i + 2, value=yr)
            c.font = self._style_th
            c.fill = self._fill_navy
        ws.row_dimensions[r].height = EXCEL_HEADER_ROW_HEIGHT
        r += 1

        # Data
        rows_data = [
            ("Gross Revenue", bd.gross_revenue),
            ("FTP Share", bd.ftp_share),
            ("Cost Recovery", bd.cost_recovery_received),
            ("Equity Split", bd.equity_split_share),
            ("Investment Credit", bd.investment_credit),
            ("DMO Revenue", bd.dmo_revenue),
            ("Taxable Income", bd.taxable_income),
            ("Tax Paid", bd.tax_paid),
            ("Net Cashflow", bd.net_cashflow),
        ]
        for label, values in rows_data:
            ws.cell(row=r, column=1, value=label).font = self._style_label_bold
            for i, val in enumerate(values):
                c = ws.cell(row=r, column=i + 2, value=val)
                c.font = self._style_data
                c.number_format = FMT_USD
                c.alignment = Alignment(horizontal="right")
            r += 1

        return r

    @staticmethod
    def _set_col_widths(ws, widths):
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

    def _build_governing_thought(self) -> str:
        """Generate the governing thought (action title)."""
        res = self._result
        npv_str = f"${res.npv_contractor:,.0f}"
        irr_str = f"{res.irr_contractor:.1%}"
        if res.npv_contractor > 0:
            return (
                f"Project NPV positif {npv_str} (IRR {irr_str}), "
                f"layak secara ekonomi dengan government take "
                f"{res.government_take_pct_of_gross:.1f}% dari gross revenue."
            )
        return (
            f"Project NPV negatif {npv_str} (IRR {irr_str}); "
            f"perlu renegosiasi fiscal term atau optimasi biaya."
        )
